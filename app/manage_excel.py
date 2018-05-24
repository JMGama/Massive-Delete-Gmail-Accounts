import openpyxl as exl
from user_interface import UserInterface

class ManageExcel:

    def __init__(self):
        """
        Load and open the Excel file with all
        the users and emails.
        """
        self.file = UserInterface.select_file()
        self.doc = exl.load_workbook(self.file)
        self.sheet = self.doc.worksheets[0]
        self.rows = self.sheet.rows
        self.current_user = 2

    def next_user(self):
        """
        Pass to the next user in the Excel file.
        """
        self.current_user += 1

    def get_name(self):
        """
        Returns the name of the user in the current row.
        """

        name = self.sheet.cell(row=self.current_user,column=1).value
        return name

    def get_email(self):
        """
        Returns the email of the user in the current row.
        """

        email = self.sheet.cell(row=self.current_user,column=2).value
        return email

    def get_total_users(self):
        """
        Returns the total of the user in the Excel file.
        """

        row_count = self.sheet.max_row
        return row_count-1 # Este -1 es debido a la fila que se ocupa unicamnete para el titulo del archivo (nombre completo, correo institucional)

    def set_unfound_user(self):
        self.sheet.cell(row=self.current_user, column=3).value = "No se encontro el usuario"
        self.doc.save(self.file)
